"""Loads an arbitrary CSV/XLSX/JSON export into plain rows (list of dict),
for the Import Wizard to preview and map - independent of what tool
produced the file or what its column names are."""

import csv
import json
import sys
from pathlib import Path

try:
    import openpyxl
    _openpyxl_import_error = None
except Exception as _exc:  # noqa: BLE001 - deliberately broad, see message below
    openpyxl = None
    _openpyxl_import_error = _exc


class UnsupportedFileError(ValueError):
    pass


def _openpyxl_missing_message() -> str:
    detail = f" (underlying error: {_openpyxl_import_error})" if _openpyxl_import_error else ""
    return (
        "Reading .xlsx requires openpyxl, and it could not be imported"
        f"{detail}.\n\n"
        f"This app is running from:\n{sys.executable}\n\n"
        "If you already ran 'pip install openpyxl', make sure you installed "
        "it into THIS interpreter specifically - e.g. run:\n"
        f'"{sys.executable}" -m pip install openpyxl\n'
        "A different 'pip install openpyxl' elsewhere (system Python, "
        "another venv, VS Code's default interpreter, etc.) won't be seen "
        "by this app if it's not the same environment."
    )


def sheet_names(path: str | Path) -> list[str]:
    """Empty list for non-xlsx files (nothing to pick)."""
    if Path(path).suffix.lower() not in (".xlsx", ".xlsm"):
        return []
    if openpyxl is None:
        raise UnsupportedFileError(_openpyxl_missing_message())
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return wb.sheetnames


def load_raw_rows(path: str | Path, sheet: str | None = None) -> list[list]:
    """Returns EVERY row (including header/junk rows) as a list of cell
    values - the wizard decides which row is the real header, since that
    varies per export (see the vCenter export with its stray junk row 2)."""
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            return [row for row in csv.reader(f)]

    if suffix in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise UnsupportedFileError(_openpyxl_missing_message())
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else v for v in row])
        return rows

    if suffix == ".json":
        raw = json.loads(path.read_text(encoding="utf-8"))
        # Accept either a plain list of objects (pvesh --output-format json
        # for a resource list), or {"data": [...]} (pvesh API wrapper shape).
        if isinstance(raw, dict) and "data" in raw:
            raw = raw["data"]
        if not isinstance(raw, list) or not raw:
            raise UnsupportedFileError("JSON file doesn't look like a list of records.")
        header = list(raw[0].keys())
        rows = [header]
        for record in raw:
            rows.append([record.get(col, "") for col in header])
        return rows

    raise UnsupportedFileError(f"Unsupported file type: {suffix or '(no extension)'}")


def rows_to_dicts(raw_rows: list[list], header_row_index: int) -> tuple[list[str], list[dict]]:
    """header_row_index is 0-based into raw_rows. Returns (header, data_rows)
    with data_rows as dicts keyed by header, skipping fully-blank rows."""
    header = [str(c).strip() if c is not None else "" for c in raw_rows[header_row_index]]
    data = []
    for raw in raw_rows[header_row_index + 1:]:
        if all(c == "" or c is None for c in raw):
            continue
        row = {header[i]: raw[i] if i < len(raw) else "" for i in range(len(header)) if header[i]}
        data.append(row)
    return header, data


def guess_header_row(raw_rows: list[list], max_scan: int = 10) -> int:
    """Best-effort guess at which row is the real header. The strongest
    signal isn't "looks like text" (data cells like '826.9 GB' are text
    too) - it's that header labels almost never contain a digit, while
    data rows almost always do. Falls back to row 0 if nothing scores
    positively - the wizard always lets the user override this anyway."""
    best_index = 0
    best_score = float("-inf")
    for i, row in enumerate(raw_rows[:max_scan]):
        non_empty = [c for c in row if c not in (None, "")]
        if not non_empty:
            continue
        text_like = sum(1 for c in non_empty if isinstance(c, str) and not _looks_numeric(c))
        digit_cells = sum(1 for c in non_empty if any(ch.isdigit() for ch in str(c)))
        blank_penalty = len(row) - len(non_empty)
        score = text_like - 3 * digit_cells - blank_penalty
        if score > best_score:
            best_score = score
            best_index = i
    return best_index


def _looks_numeric(text: str) -> bool:
    try:
        float(text.replace(",", ""))
        return True
    except ValueError:
        return False
