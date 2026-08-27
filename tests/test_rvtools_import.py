import openpyxl
import pytest

from src.persistence import rvtools_import


def _write_sample_export(path, include_vhost=True, include_vinfo=True):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if include_vhost:
        vhost = wb.create_sheet("vHost")
        vhost.append(["Host", "Datacenter", "Cluster", "# CPU", "Cores per CPU", "# Cores", "# Memory", "CPU Model"])
        vhost.append(["esxi-01.lab.local", "DC1", "Cluster1", 2, 24, 48, 524288, "Intel Xeon Gold 6338"])
        vhost.append(["esxi-02.lab.local", "DC1", "Cluster1", 2, 24, 48, 524288, "Intel Xeon Gold 6338"])

    if include_vinfo:
        vinfo = wb.create_sheet("vInfo")
        vinfo.append(["VM", "Powerstate", "Template", "CPUs", "Memory", "Provisioned MiB", "Host", "Cluster", "Datacenter"])
        vinfo.append(["app-01", "poweredOn", False, 4, 16384, 153600, "esxi-01.lab.local", "Cluster1", "DC1"])
        vinfo.append(["db-01", "poweredOn", False, 8, 65536, 512000, "esxi-01.lab.local", "Cluster1", "DC1"])
        vinfo.append(["old-test", "poweredOff", False, 2, 4096, 51200, "esxi-02.lab.local", "Cluster1", "DC1"])

    wb.save(path)


def test_preview_counts(tmp_path):
    path = tmp_path / "export.xlsx"
    _write_sample_export(path)
    vm_count, host_count = rvtools_import.preview_counts(path)
    assert vm_count == 3
    assert host_count == 2


def test_import_servers_maps_fields_and_converts_mib_to_gib(tmp_path):
    path = tmp_path / "export.xlsx"
    _write_sample_export(path)
    servers = rvtools_import.import_servers(path)

    assert len(servers) == 2
    assert servers[0].name == "esxi-01.lab.local"
    assert servers[0].sockets == 2
    assert servers[0].cores_per_socket == 24
    assert servers[0].ram_gb == 512  # 524288 MiB / 1024
    assert servers[0].cpu_model == "Intel Xeon Gold 6338"
    assert servers[0].hyperthreading_enabled is False  # fixture has no HT Available/HT Active columns - falls back to the conservative default


def test_import_vms_maps_fields_and_converts_mib_to_gib(tmp_path):
    path = tmp_path / "export.xlsx"
    _write_sample_export(path)
    vms = rvtools_import.import_vms(path)

    assert len(vms) == 3
    assert vms[0].name == "app-01"
    assert vms[0].vcpu == 4
    assert vms[0].ram_gb == 16
    assert vms[0].disk_gb == 150
    assert vms[0].powered_on is True
    assert vms[2].powered_on is False  # old-test is poweredOff


def test_non_rvtools_file_raises_clear_error(tmp_path):
    path = tmp_path / "unrelated.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append(["Name", "Age"])
    wb.active.append(["Bob", 30])
    wb.save(path)

    with pytest.raises(rvtools_import.RVToolsImportError):
        rvtools_import.preview_counts(path)


def test_missing_vinfo_sheet_returns_zero_vms(tmp_path):
    path = tmp_path / "host_only.xlsx"
    _write_sample_export(path, include_vinfo=False)

    servers = rvtools_import.import_servers(path)
    vms = rvtools_import.import_vms(path)
    assert len(servers) == 2
    assert len(vms) == 0


def test_missing_vhost_sheet_returns_zero_servers(tmp_path):
    path = tmp_path / "vms_only.xlsx"
    _write_sample_export(path, include_vhost=False)

    servers = rvtools_import.import_servers(path)
    vms = rvtools_import.import_vms(path)
    assert len(servers) == 0
    assert len(vms) == 3


def test_import_servers_detects_ip_address_when_host_is_an_ip(tmp_path):
    """RVTools' vHost sheet sometimes identifies hosts by IP directly -
    when it does, populate ip_address too, not just name."""
    path = tmp_path / "export.xlsx"
    _write_sample_export(path)
    servers = rvtools_import.import_servers(path)
    assert servers[0].ip_address == "esxi-01.lab.local" or servers[0].ip_address == ""
    # the synthetic fixture uses hostnames, not IPs - confirm hostname does NOT get miscategorized as an IP
    assert servers[0].ip_address == ""


def test_import_servers_detects_ip_address_from_real_ip_host(tmp_path):
    path = tmp_path / "export_ip.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vhost = wb.create_sheet("vHost")
    vhost.append(["Host", "# CPU", "Cores per CPU", "# Memory"])
    vhost.append(["10.88.1.10", 2, 24, 524288])
    wb.save(path)

    servers = rvtools_import.import_servers(path)
    assert servers[0].name == "10.88.1.10"
    assert servers[0].ip_address == "10.88.1.10"


def test_import_vms_populates_ip_address_when_present(tmp_path):
    path = tmp_path / "export_vm_ip.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vinfo = wb.create_sheet("vInfo")
    vinfo.append(["VM", "Powerstate", "CPUs", "Memory", "Total disk capacity MiB", "Primary IP Address"])
    vinfo.append(["app-01", "poweredOn", 4, 16384, 153600, "10.20.1.15"])
    vinfo.append(["app-02", "poweredOn", 2, 8192, 51200, ""])
    wb.save(path)

    vms = rvtools_import.import_vms(path)

    assert vms[0].ip_address == "10.20.1.15"
    assert vms[1].ip_address == ""


def test_import_servers_detects_ht_active(tmp_path):
    """Confirmed against a real RVTools export: HT Available/HT Active
    are reliably present and correct - a previous version of this
    importer incorrectly assumed otherwise and always defaulted to HT
    off with threads_per_core=1, which meant toggling the Hyperthreading
    checkbox after import had literally no effect (1 thread/core makes
    the HT-enabled/disabled multiplication a no-op either way)."""
    path = tmp_path / "ht_on.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vhost = wb.create_sheet("vHost")
    vhost.append(["Host", "# CPU", "Cores per CPU", "# Memory", "HT Available", "HT Active"])
    vhost.append(["esxi-01", 2, 10, 262144, True, True])
    wb.save(path)

    servers = rvtools_import.import_servers(path)

    assert servers[0].hyperthreading_enabled is True
    assert servers[0].threads_per_core == 2
    assert servers[0].effective_cores == 40  # 2 sockets x 10 cores x 2 threads


def test_import_servers_ht_available_but_not_active_keeps_thread_width(tmp_path):
    """HT capable but disabled (e.g. in BIOS) - threads_per_core stays 2
    (the real SMT width) even though it's currently off, so toggling
    Hyperthreading back on in the app actually does something instead
    of being stuck at 1 thread/core forever."""
    path = tmp_path / "ht_off.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vhost = wb.create_sheet("vHost")
    vhost.append(["Host", "# CPU", "Cores per CPU", "# Memory", "HT Available", "HT Active"])
    vhost.append(["esxi-01", 2, 10, 262144, True, False])
    wb.save(path)

    servers = rvtools_import.import_servers(path)

    assert servers[0].hyperthreading_enabled is False
    assert servers[0].threads_per_core == 2
    assert servers[0].effective_cores == 20  # HT off right now: no doubling

    # Confirm toggling it back on actually works (the whole point of the fix)
    servers[0].hyperthreading_enabled = True
    assert servers[0].effective_cores == 40


def test_import_servers_no_ht_capability(tmp_path):
    path = tmp_path / "no_ht.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vhost = wb.create_sheet("vHost")
    vhost.append(["Host", "# CPU", "Cores per CPU", "# Memory", "HT Available", "HT Active"])
    vhost.append(["esxi-01", 2, 10, 262144, False, False])
    wb.save(path)

    servers = rvtools_import.import_servers(path)

    assert servers[0].hyperthreading_enabled is False
    assert servers[0].threads_per_core == 1
    assert servers[0].effective_cores == 20


def test_detect_datacenters_single_value(tmp_path):
    path = tmp_path / "single_dc.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vinfo = wb.create_sheet("vInfo")
    vinfo.append(["VM", "Powerstate", "CPUs", "Memory", "Datacenter"])
    vinfo.append(["app-01", "poweredOn", 4, 16384, "DC1"])
    vinfo.append(["app-02", "poweredOn", 2, 8192, "DC1"])
    wb.save(path)

    dcs = rvtools_import.detect_datacenters(path)
    assert dcs == ["DC1"]


def test_detect_datacenters_multiple_values(tmp_path):
    path = tmp_path / "multi_dc.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vinfo = wb.create_sheet("vInfo")
    vinfo.append(["VM", "Powerstate", "CPUs", "Memory", "Datacenter"])
    vinfo.append(["app-01", "poweredOn", 4, 16384, "DC-Primary"])
    vinfo.append(["app-02", "poweredOn", 2, 8192, "DC-DR"])
    wb.save(path)

    dcs = rvtools_import.detect_datacenters(path)
    assert dcs == ["DC-DR", "DC-Primary"]  # sorted


def test_detect_datacenters_no_column_returns_empty(tmp_path):
    path = tmp_path / "no_dc_col.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vinfo = wb.create_sheet("vInfo")
    vinfo.append(["VM", "Powerstate", "CPUs", "Memory"])
    vinfo.append(["app-01", "poweredOn", 4, 16384])
    wb.save(path)

    assert rvtools_import.detect_datacenters(path) == []


def test_import_servers_picks_up_cluster_name(tmp_path):
    path = tmp_path / "cluster.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vhost = wb.create_sheet("vHost")
    vhost.append(["Host", "# CPU", "Cores per CPU", "# Memory", "Cluster"])
    vhost.append(["esxi-01", 2, 10, 262144, "vSAN_HPM"])
    wb.save(path)

    servers = rvtools_import.import_servers(path)
    assert servers[0].cluster_name == "vSAN_HPM"


def test_site_map_routes_servers_and_vms_by_datacenter(tmp_path):
    path = tmp_path / "sites.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vhost = wb.create_sheet("vHost")
    vhost.append(["Host", "# CPU", "Cores per CPU", "# Memory", "Datacenter"])
    vhost.append(["esxi-p01", 2, 10, 262144, "DC-Primary"])
    vhost.append(["esxi-dr01", 2, 10, 262144, "DC-DR"])
    vinfo = wb.create_sheet("vInfo")
    vinfo.append(["VM", "Powerstate", "CPUs", "Memory", "Datacenter"])
    vinfo.append(["app-01", "poweredOn", 4, 16384, "DC-Primary"])
    vinfo.append(["dc-01", "poweredOn", 2, 8192, "DC-DR"])
    wb.save(path)

    site_map = {"DC-Primary": "Primary", "DC-DR": "DR"}
    servers = rvtools_import.import_servers(path, site_map=site_map)
    vms = rvtools_import.import_vms(path, site_map=site_map)

    assert {s.name: s.site for s in servers} == {"esxi-p01": "Primary", "esxi-dr01": "DR"}
    assert {v.name: v.site for v in vms} == {"app-01": "Primary", "dc-01": "DR"}


def test_site_map_falls_back_to_default_site_for_unmapped_datacenter(tmp_path):
    """An unexpected/unmapped Datacenter value must not crash - falls
    back to the caller's default site instead."""
    path = tmp_path / "unmapped.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vhost = wb.create_sheet("vHost")
    vhost.append(["Host", "# CPU", "Cores per CPU", "# Memory", "Datacenter"])
    vhost.append(["esxi-01", 2, 10, 262144, "SomeOtherDC"])
    wb.save(path)

    servers = rvtools_import.import_servers(path, site="Primary", site_map={"DC-Primary": "Primary", "DC-DR": "DR"})
    assert servers[0].site == "Primary"


def test_os_preference_config_file_falls_back_to_tools_when_blank(tmp_path):
    path = tmp_path / "os_fallback.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vinfo = wb.create_sheet("vInfo")
    vinfo.append([
        "VM", "Powerstate", "CPUs", "Memory",
        "OS according to the configuration file", "OS according to the VMware Tools",
    ])
    vinfo.append(["app-01", "poweredOn", 4, 16384, "", "Ubuntu Linux (64-bit)"])
    wb.save(path)

    vms = rvtools_import.import_vms(path, os_preference="config")
    assert vms[0].os == "Ubuntu Linux (64-bit)"  # config was blank, fell back to tools


def test_os_preference_prefers_the_requested_source_when_both_present(tmp_path):
    path = tmp_path / "os_both.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vinfo = wb.create_sheet("vInfo")
    vinfo.append([
        "VM", "Powerstate", "CPUs", "Memory",
        "OS according to the configuration file", "OS according to the VMware Tools",
    ])
    vinfo.append(["app-01", "poweredOn", 4, 16384, "Windows Server 2012 (64-bit)", "Windows Server 2016 or later (64-bit)"])
    wb.save(path)

    vms_config = rvtools_import.import_vms(path, os_preference="config")
    vms_tools = rvtools_import.import_vms(path, os_preference="tools")

    assert vms_config[0].os == "Windows Server 2012 (64-bit)"
    assert vms_tools[0].os == "Windows Server 2016 or later (64-bit)"


def test_import_switches_deduplicates_by_name(tmp_path):
    path = tmp_path / "switches.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vswitch = wb.create_sheet("vSwitch")
    vswitch.append(["Host", "Datacenter", "Cluster", "Switch"])
    vswitch.append(["esxi-01", "DC1", "Cluster1", "vSwitch0"])
    vswitch.append(["esxi-02", "DC1", "Cluster1", "vSwitch0"])  # same switch, different host
    vswitch.append(["esxi-01", "DC1", "Cluster1", "vSwitchBMC"])
    wb.save(path)

    switches = rvtools_import.import_switches(path)

    assert len(switches) == 2
    names = {sw.name for sw in switches}
    assert names == {"vSwitch0", "vSwitchBMC"}


def test_import_switches_missing_sheet_returns_empty(tmp_path):
    path = tmp_path / "no_switches.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vinfo = wb.create_sheet("vInfo")
    vinfo.append(["VM", "Powerstate", "CPUs", "Memory"])
    vinfo.append(["app-01", "poweredOn", 4, 16384])
    wb.save(path)

    assert rvtools_import.import_switches(path) == []


def test_round_up_to_known_ram_gb_matches_the_real_discrepancy_found():
    """The exact case that motivated this: a real export reported
    383.7GB for hosts confirmed to have 256GB of DIMMs installed -
    round up to the nearest standard config (384GB) instead of keeping
    the odd exact figure."""
    assert rvtools_import.round_up_to_known_ram_gb(383.68) == 384


def test_round_up_to_known_ram_gb_exact_match_stays_unchanged():
    assert rvtools_import.round_up_to_known_ram_gb(256.0) == 256


def test_round_up_to_known_ram_gb_rounds_a_nonstandard_value_up():
    assert rvtools_import.round_up_to_known_ram_gb(200) == 256


def test_round_up_to_known_ram_gb_beyond_known_list_keeps_reported_value():
    assert rvtools_import.round_up_to_known_ram_gb(10000) == 10000


def test_import_servers_rounds_ram_to_known_config(tmp_path):
    path = tmp_path / "ram_round.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vhost = wb.create_sheet("vHost")
    vhost.append(["Host", "# CPU", "Cores per CPU", "# Memory"])
    vhost.append(["esxi-01", 2, 16, 392888])  # 383.68 GiB raw
    wb.save(path)

    servers = rvtools_import.import_servers(path)

    assert servers[0].ram_gb == 384


def test_import_vms_ram_is_not_rounded_to_known_server_configs(tmp_path):
    """VMs aren't built from physical DIMMs - a VM can legitimately have
    any RAM allocation, so this rounding must only apply to Server
    imports, never VM imports."""
    path = tmp_path / "vm_ram.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vinfo = wb.create_sheet("vInfo")
    vinfo.append(["VM", "Powerstate", "CPUs", "Memory"])
    vinfo.append(["app-01", "poweredOn", 2, 6144])  # 6 GiB - not a "known server RAM" value
    wb.save(path)

    vms = rvtools_import.import_vms(path)

    assert vms[0].ram_gb == 6  # exact, untouched
