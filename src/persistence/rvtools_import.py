"""Parses a standard RVTools export (.xlsx, multiple sheets - "Export all
to Excel" from the RVTools UI, the common case) into Server,
VirtualMachine, and NetworkSwitch objects.

Three sheets are used: "vHost" (one row per physical host -> Server),
"vInfo" (one row per VM -> VirtualMachine), and "vSwitch" (one row per
host-switch pairing -> NetworkSwitch, deduplicated by name).

Site assignment: RVTools doesn't have a Primary/DR concept of its own,
but its "Datacenter" column often does distinguish sites for real
multi-site environments (Primary and DR living in the same vCenter as
two Datacenter objects). If the export has just ONE Datacenter value,
everything goes to whichever site the caller passes in `site` - same
as before. If it has MORE than one, the caller can pass `site_map`
(Datacenter value -> "Primary"/"DR") to route each row to the right
site individually; `detect_datacenters()` is what a dialog calls first
to decide whether that mapping step is even needed.

Column names have drifted a little across RVTools versions, so lookups
try a short list of known aliases per field rather than one exact name.
RVTools' own "MB"/"GB" labels are actually MiB/GiB (base-2) - see
https://sizing-workshop.readthedocs.io/en/latest/datacollection/rvtools/rvtools.html -
so every memory/disk value is divided by 1024, not 1000.
"""

import openpyxl

from src.models.server import Server
from src.models.virtual_machine import VirtualMachine
from src.models.network_switch import NetworkSwitch

MIB_PER_GIB = 1024

# Real servers are built from a short, well-known list of standard RAM
# configurations (driven by DIMM slot count x DIMM size) - a host never
# actually has, say, 383GB installed. RVTools/vCenter's reported host
# memory can include some overhead beyond what's physically installed
# (confirmed on a real export: 383.7GB reported for hosts the customer
# confirmed have 256GB of DIMMs - a ~127GB gap neither "Memory Tiering"
# (checked: off) nor anything else in the export could explain). Round
# UP to the nearest known configuration instead of keeping the odd
# exact figure - the nearest at-or-above value, not nearest-either-way,
# since a host can't have LESS installed than what vCenter measured.
KNOWN_SERVER_RAM_GB = [
    8, 16, 32, 64, 96, 128, 192, 256, 384, 512, 768,
    1024, 1536, 2048, 3072, 4096, 6144, 8192,
]


def round_up_to_known_ram_gb(raw_gb: float) -> int:
    for known in KNOWN_SERVER_RAM_GB:
        if known >= raw_gb:
            return known
    return round(raw_gb)  # bigger than the whole known list - keep as reported rather than guess further


class RVToolsImportError(ValueError):
    pass


def _normalize(header) -> str:
    return str(header or "").strip().lower()


def _find_column(headers: dict[str, int], aliases: list[str]) -> int | None:
    for alias in aliases:
        idx = headers.get(_normalize(alias))
        if idx is not None:
            return idx
    return None


def _header_index(sheet) -> dict[str, int]:
    headers = {}
    for col_idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1))):
        if cell.value:
            headers[_normalize(cell.value)] = col_idx
    return headers


def _find_sheet(workbook, *candidates: str):
    for name in workbook.sheetnames:
        if _normalize(name) in {_normalize(c) for c in candidates}:
            return workbook[name]
    return None


def _cell(row, idx: int | None):
    if idx is None or idx >= len(row):
        return None
    return row[idx].value


def _looks_like_ipv4(text: str) -> bool:
    parts = text.strip().split(".")
    if len(parts) != 4:
        return False
    return all(part.isdigit() and 0 <= int(part) <= 255 for part in parts)


def _resolve_site(datacenter_value, site: str, site_map: dict[str, str] | None) -> str:
    """Falls back to `site` when there's no map, no Datacenter column in
    this sheet, or the value found isn't a key in the map - never
    crashes on an unexpected Datacenter value, just uses the default."""
    if not site_map:
        return site
    key = str(datacenter_value or "").strip()
    return site_map.get(key, site)


def preview_counts(path) -> tuple[int, int]:
    """Returns (vm_count, host_count) without building full objects -
    for the import dialog to show "will import X VMs and Y hosts" before
    the user commits."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        vinfo = _find_sheet(workbook, "vInfo", "tabvInfo")
        vhost = _find_sheet(workbook, "vHost", "tabvHost")
        if vinfo is None and vhost is None:
            raise RVToolsImportError(
                "This doesn't look like an RVTools export - no 'vInfo' or "
                "'vHost' sheet found. Use RVTools' File > Export all to "
                "Excel (not a single-tab export)."
            )
        vm_count = max(0, (vinfo.max_row or 1) - 1) if vinfo else 0
        host_count = max(0, (vhost.max_row or 1) - 1) if vhost else 0
        return vm_count, host_count
    finally:
        workbook.close()


def detect_datacenters(path) -> list[str]:
    """Distinct, non-blank Datacenter values found across vInfo (falls
    back to vHost if vInfo is missing) - for the import dialog to decide
    whether a Primary/DR mapping step is needed at all. A single value
    (the overwhelmingly common case) means no mapping is needed; the
    caller just picks one target site like before."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _find_sheet(workbook, "vInfo", "tabvInfo") or _find_sheet(workbook, "vHost", "tabvHost")
        if sheet is None:
            return []

        headers = _header_index(sheet)
        col_dc = _find_column(headers, ["Datacenter"])
        if col_dc is None:
            return []

        found = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            value = row[col_dc] if col_dc < len(row) else None
            if value:
                found.add(str(value).strip())
        return sorted(found)
    finally:
        workbook.close()


def import_servers(
    path, site: str = "Primary", site_map: dict[str, str] | None = None,
) -> list[Server]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _find_sheet(workbook, "vHost", "tabvHost")
        if sheet is None:
            return []

        headers = _header_index(sheet)
        col_host = _find_column(headers, ["Host"])
        col_sockets = _find_column(headers, ["# CPU", "#CPU", "CPUs", "Sockets"])
        col_cores_per_cpu = _find_column(headers, ["Cores per CPU", "Cores p/s"])
        col_ram_mib = _find_column(headers, ["# Memory", "Memory"])
        col_cpu_model = _find_column(headers, ["CPU Model"])
        col_ht_active = _find_column(headers, ["HT Active"])
        col_ht_available = _find_column(headers, ["HT Available"])
        col_cluster = _find_column(headers, ["Cluster"])
        col_datacenter = _find_column(headers, ["Datacenter"])

        if col_host is None:
            raise RVToolsImportError(
                "The 'vHost' sheet is missing a 'Host' column - this doesn't "
                "look like a standard RVTools export."
            )

        servers = []
        for row in sheet.iter_rows(min_row=2, values_only=False):
            host_name = _cell(row, col_host)
            if not host_name:
                continue

            server = Server.create_default()
            server.name = str(host_name)
            server.site = _resolve_site(_cell(row, col_datacenter), site, site_map)
            if _looks_like_ipv4(str(host_name)):
                server.ip_address = str(host_name)
            server.sockets = int(_cell(row, col_sockets) or 1)
            server.cores_per_socket = int(_cell(row, col_cores_per_cpu) or 1)
            ht_active = _cell(row, col_ht_active)
            ht_available = _cell(row, col_ht_available)
            if ht_active is not None or ht_available is not None:
                # RVTools reliably reports this (confirmed against a real
                # export) - HT Available means the CPU is SMT-capable at
                # all; HT Active means it's actually turned on right now.
                # Keeping threads_per_core=2 when merely available-but-off
                # matches this app's own model: toggling the Hyperthreading
                # checkbox later shouldn't lose the host's real SMT width.
                is_capable = bool(ht_available) or bool(ht_active)
                server.threads_per_core = 2 if is_capable else 1
                server.hyperthreading_enabled = bool(ht_active)
            else:
                # Neither column present in this export - fall back to the
                # old conservative default rather than guessing.
                server.threads_per_core = 1
                server.hyperthreading_enabled = False
            ram_mib = _cell(row, col_ram_mib)
            server.ram_gb = round_up_to_known_ram_gb(float(ram_mib) / MIB_PER_GIB) if ram_mib else 0
            cpu_model = _cell(row, col_cpu_model)
            if cpu_model:
                server.cpu_model = str(cpu_model)
            cluster_name = _cell(row, col_cluster)
            if cluster_name:
                server.cluster_name = str(cluster_name)
            server.notes = "Imported from RVTools - review vendor/model and warranty manually."
            servers.append(server)

        return servers
    finally:
        workbook.close()


def import_vms(
    path, site: str = "Primary", site_map: dict[str, str] | None = None,
    os_preference: str = "config",
) -> list[VirtualMachine]:
    """os_preference: "config" prefers "OS according to the configuration
    file" (always populated, but only as accurate as what was declared
    when the VM was created); "tools" prefers "OS according to the
    VMware Tools" (detected live, but blank if Tools isn't installed/
    running). Either way, falls back to the other column when the
    preferred one is blank for a given VM."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _find_sheet(workbook, "vInfo", "tabvInfo")
        if sheet is None:
            return []

        headers = _header_index(sheet)
        col_name = _find_column(headers, ["VM"])
        col_power = _find_column(headers, ["Powerstate"])
        col_vcpu = _find_column(headers, ["CPUs"])
        col_ram_mib = _find_column(headers, ["Memory"])
        col_disk_mib = _find_column(headers, [
            "Total disk capacity MiB", "Total disk capacity MB",
            "Provisioned MiB", "Provisioned MB", "Provisioned in MB", "In Use MiB", "In Use MB",
        ])
        col_ip = _find_column(headers, ["Primary IP Address", "IP Address"])
        col_os_config = _find_column(headers, ["OS according to the configuration file"])
        col_os_tools = _find_column(headers, ["OS according to the VMware Tools"])
        col_datacenter = _find_column(headers, ["Datacenter"])

        if col_name is None:
            raise RVToolsImportError(
                "The 'vInfo' sheet is missing a 'VM' column - this doesn't "
                "look like a standard RVTools export."
            )

        preferred_os_col, fallback_os_col = (
            (col_os_tools, col_os_config) if os_preference == "tools" else (col_os_config, col_os_tools)
        )

        vms = []
        for row in sheet.iter_rows(min_row=2, values_only=False):
            vm_name = _cell(row, col_name)
            if not vm_name:
                continue

            vm = VirtualMachine.create_default()
            vm.name = str(vm_name)
            vm.site = _resolve_site(_cell(row, col_datacenter), site, site_map)
            vm.vcpu = int(_cell(row, col_vcpu) or 0)
            ram_mib = _cell(row, col_ram_mib)
            vm.ram_gb = int(float(ram_mib) / MIB_PER_GIB) if ram_mib else 0
            disk_mib = _cell(row, col_disk_mib)
            vm.disk_gb = int(float(disk_mib) / MIB_PER_GIB) if disk_mib else 0
            power = _cell(row, col_power)
            vm.powered_on = _normalize(power) == "poweredon" if power else True
            ip_address = _cell(row, col_ip)
            if ip_address:
                vm.ip_address = str(ip_address)
            os_value = _cell(row, preferred_os_col) or _cell(row, fallback_os_col)
            if os_value:
                vm.os = str(os_value)
            vm.notes = "Imported from RVTools - review Workload Tier and DR Protected manually."
            vms.append(vm)

        return vms
    finally:
        workbook.close()


def import_switches(
    path, site: str = "Primary", site_map: dict[str, str] | None = None,
) -> list[NetworkSwitch]:
    """One NetworkSwitch per distinct Switch name found on the vSwitch
    sheet - name only, nothing else (port counts, speed, etc. are set
    manually afterward - RVTools' switch data doesn't map cleanly onto
    this app's port-inventory fields). Deduplicated: the same switch
    name commonly appears once per host that connects to it."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _find_sheet(workbook, "vSwitch", "tabvSwitch")
        if sheet is None:
            return []

        headers = _header_index(sheet)
        col_switch = _find_column(headers, ["Switch"])
        col_datacenter = _find_column(headers, ["Datacenter"])

        if col_switch is None:
            return []

        seen: dict[str, NetworkSwitch] = {}
        for row in sheet.iter_rows(min_row=2, values_only=False):
            switch_name = _cell(row, col_switch)
            if not switch_name:
                continue
            switch_name = str(switch_name)
            if switch_name in seen:
                continue

            switch = NetworkSwitch.create_default()
            switch.name = switch_name
            switch.site = _resolve_site(_cell(row, col_datacenter), site, site_map)
            switch.notes = "Imported from RVTools - name only, review port counts/speed manually."
            seen[switch_name] = switch

        return list(seen.values())
    finally:
        workbook.close()
