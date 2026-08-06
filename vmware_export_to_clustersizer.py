#!/usr/bin/env python3
"""
Converts a VMware vCenter "VMs for Cluster" .xlsx export into a CSV that
ClusterSizer's VMs tab can import directly.

Usage:
    python vmware_export_to_clustersizer.py INPUT.xlsx [OUTPUT.csv]
                                             [--site Primary|DR]
                                             [--include-vcls]

Example:
    python vmware_export_to_clustersizer.py HPM_vSAN_stanje_4_8_2026.xlsx vms.csv

What it does:
    - Reads the "Name / State / Status / Provisioned Space / Used Space /
      Host CPU / Host Mem / Guest OS / Memory Size / CPUs / VMware Tools..."
      sheet that vCenter produces for a cluster VM list.
    - Skips vCenter's known malformed second header row (a "Status"" +
      "DNS Name" spillover row that some vCenter/vSphere Client export
      versions produce - harmless, just needs to be skipped).
    - Maps CPUs -> vcpu, Memory Size -> ram_gb, Provisioned Space -> disk_gb
      (provisioned/allocated, not "Used Space" - capacity planning cares
      about what's allocated, not what's currently written to disk).
    - State "Powered On"/"Powered Off" -> powered_on True/False.
    - By default EXCLUDES vCLS-* system VMs (vSphere Cluster Services -
      internal VMware infrastructure VMs, not real workloads you'd size
      capacity for). Use --include-vcls to keep them.
    - dr_protected/dr_vcpu/dr_ram_gb/dr_disk_gb are NOT in a vCenter export
      (vCenter doesn't know your DR replication setup) - they're written
      with dr_protected=False and dr_* mirroring the primary values, so
      the fields are valid but inert until you flag specific VMs as
      DR Protected inside ClusterSizer itself.
    - Guest OS and DNS Name go into the notes column for reference.
"""

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("This script needs openpyxl: pip install openpyxl")


# Must match VM_FIELDS in ClusterSizer's src/persistence/csv_io.py exactly -
# the import is strictly schema-checked, so this header can't drift.
VM_CSV_FIELDS = [
    "name", "site", "vcpu", "ram_gb", "disk_gb", "powered_on",
    "dr_protected", "dr_vcpu", "dr_ram_gb", "dr_disk_gb", "notes",
]

SIZE_UNITS_TO_GB = {
    "B": 1 / (1024 ** 3),
    "KB": 1 / (1024 ** 2),
    "MB": 1 / 1024,
    "GB": 1.0,
    "TB": 1024.0,
}


def parse_size_to_gb(text: str | None) -> float:
    """Parses vCenter size strings like '826.9 GB', '4.32 TB', '160 MB',
    '1,009.86 GB', '0 B' into a GB float."""
    if not text:
        return 0.0
    text = str(text).strip().replace(",", "")
    match = re.match(r"^([\d.]+)\s*([A-Za-z]+)$", text)
    if not match:
        return 0.0
    value, unit = match.groups()
    unit = unit.upper()
    factor = SIZE_UNITS_TO_GB.get(unit)
    if factor is None:
        return 0.0
    return round(float(value) * factor, 2)


def parse_vcpu(value) -> int:
    try:
        return max(1, int(round(float(value))))
    except (TypeError, ValueError):
        return 1


def parse_powered_on(state: str | None) -> bool:
    return str(state or "").strip().lower() == "powered on"


def read_vcenter_export(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active  # the export has a single sheet

    header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col = {name: header.index(name) for name in [
            "Name", "State", "Provisioned Space", "Guest OS", "Memory Size", "CPUs",
        ]}
    except ValueError as exc:
        sys.exit(
            f"Expected column not found in the export header ({exc}). "
            f"Header row read was: {header}\n"
            "This script expects vCenter's standard 'VMs for Cluster' export "
            "columns - if vCenter changed its export format, the column "
            "mapping near the top of this script needs updating."
        )

    # DNS Name is typically the last column but has no clean header of its
    # own in this export (see module docstring) - grab it positionally.
    dns_col = len(header) - 1

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[col["Name"]]
        if not name or str(name).strip() in ("Status\"",):
            continue  # skip the known malformed spillover row, and blanks
        rows.append({
            "name": str(name).strip(),
            "state": row[col["State"]],
            "provisioned": row[col["Provisioned Space"]],
            "guest_os": row[col["Guest OS"]],
            "memory_size": row[col["Memory Size"]],
            "cpus": row[col["CPUs"]],
            "dns_name": row[dns_col] if dns_col < len(row) else None,
        })
    return rows


def convert(rows: list[dict], site: str, include_vcls: bool) -> tuple[list[dict], int]:
    out_rows = []
    excluded = 0

    for r in rows:
        if not include_vcls and r["name"].startswith("vCLS-"):
            excluded += 1
            continue

        vcpu = parse_vcpu(r["cpus"])
        ram_gb = parse_size_to_gb(r["memory_size"])
        disk_gb = parse_size_to_gb(r["provisioned"])

        notes_parts = []
        if r["guest_os"]:
            notes_parts.append(str(r["guest_os"]))
        if r["dns_name"]:
            notes_parts.append(f"DNS: {r['dns_name']}")
        notes = " | ".join(notes_parts)

        out_rows.append({
            "name": r["name"],
            "site": site,
            "vcpu": vcpu,
            "ram_gb": ram_gb,
            "disk_gb": disk_gb,
            "powered_on": parse_powered_on(r["state"]),
            "dr_protected": False,
            "dr_vcpu": vcpu,
            "dr_ram_gb": ram_gb,
            "dr_disk_gb": disk_gb,
            "notes": notes,
        })

    return out_rows, excluded


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input", type=Path, help="vCenter .xlsx export")
    parser.add_argument("output", type=Path, nargs="?", default=None, help="output CSV (default: vms.csv next to input)")
    parser.add_argument("--site", choices=["Primary", "DR"], default="Primary", help="ClusterSizer site to tag every VM with (default: Primary)")
    parser.add_argument("--include-vcls", action="store_true", help="keep vCLS-* system VMs (excluded by default)")
    args = parser.parse_args()

    if not args.input.exists():
        sys.exit(f"File not found: {args.input}")

    output_path = args.output or args.input.with_name("vms.csv")

    raw_rows = read_vcenter_export(args.input)
    converted, excluded = convert(raw_rows, site=args.site, include_vcls=args.include_vcls)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=VM_CSV_FIELDS)
        writer.writeheader()
        writer.writerows(converted)

    powered_off = sum(1 for r in converted if not r["powered_on"])
    print(f"Converted {len(converted)} VM(s) -> {output_path}")
    if excluded:
        print(f"Excluded {excluded} vCLS system VM(s) (use --include-vcls to keep them)")
    print(f"Of those, {powered_off} are currently Powered Off (still imported, just flagged)")
    print(f"All VMs tagged as site={args.site} - edit in ClusterSizer, or re-run with --site DR for a DR export.")
    print("dr_protected is False for everyone by default - flag DR-replicated VMs inside ClusterSizer itself.")


if __name__ == "__main__":
    main()
